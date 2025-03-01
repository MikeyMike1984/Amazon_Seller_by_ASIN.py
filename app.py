import os
import httpx
import ssl
import asyncio
from bs4 import BeautifulSoup
from datetime import datetime
import logging
import csv
import random
import io
from flask import Flask, request, jsonify, render_template, make_response
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from tenacity import retry, stop_after_attempt, wait_exponential

# Logging setup
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(message)s')

# Flask app setup
app = Flask(__name__, static_url_path='/static')
app.config['UPLOAD_FOLDER'] = 'uploads'
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Constants
ASINS = []  # Updated dynamically
BASE_URL = "https://www.amazon.com/gp/aod/ajax/ref=dp_aod_NEW_mbc?asin={}"
SEMAPHORE_LIMIT = 3
BATCH_SIZE = 10

# Headers to mimic a real browser
HEADERS = [
    {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
        "Accept-Language": "en-US,en;q=0.9",
        "Accept-Encoding": "gzip, deflate, br",
        "Connection": "keep-alive",
        "DNT": "1",
        "Upgrade-Insecure-Requests": "1",
    },
    {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.212 Safari/537.36",
        "Accept-Language": "en-US,en;q=0.9",
        "Accept-Encoding": "gzip, deflate, br",
        "Connection": "keep-alive",
        "DNT": "1",
    },
]

# Configure SSL context
ssl_context = ssl.create_default_context()
ssl_context.minimum_version = ssl.TLSVersion.TLSv1_2
ssl_context.maximum_version = ssl.TLSVersion.TLSv1_3

# In-memory cache
cache = {}

def create_csv_string(data):
    """
    Creates a CSV string from the scraped data.
    """
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write headers
    writer.writerow(["ASIN", "Seller", "Date", "Time"])
    
    # Write data
    writer.writerows(data)
    
    return output.getvalue()

@retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=1, min=2, max=10))
async def fetch_and_parse(asin, client):
    if asin in cache:
        return cache[asin]

    try:
        url = BASE_URL.format(asin)
        headers = random.choice(HEADERS)
        response = await client.get(url, headers=headers)
        response.raise_for_status()

        soup = BeautifulSoup(response.text, 'html.parser')
        sellers = soup.select('#aod-offer-soldBy .a-fixed-left-grid-col.a-col-right a')

        unique_sellers = list(set(seller.text.strip() for seller in sellers if seller.text.strip()))
        timestamp = datetime.now()
        formatted_data = [
            [asin, seller, timestamp.strftime("%Y-%m-%d"), timestamp.strftime("%H:%M:%S")]
            for seller in unique_sellers
        ]

        cache[asin] = formatted_data
        return formatted_data
    except Exception as e:
        logging.error(f"Error scraping ASIN {asin}: {e}")
        return []

async def fetch_with_semaphore(asin, client, semaphore):
    async with semaphore:
        return await fetch_and_parse(asin, client)

async def scrape_sellers():
    semaphore = asyncio.Semaphore(SEMAPHORE_LIMIT)
    async with httpx.AsyncClient(http2=True, verify=ssl_context) as client:
        tasks = [fetch_with_semaphore(asin, client, semaphore) for asin in ASINS]
        results = await asyncio.gather(*tasks)

    all_seller_data = [row for result in results for row in result]
    return all_seller_data

def extract_asins_from_file(file_path):
    """
    Extracts ASINs from the uploaded .xlsx file.
    Assumes that ASINs are located in column A of the first worksheet.
    """
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active  # Get the first worksheet
        asins = []

        # Assuming ASINs are in column A, starting from the first row
        for row in sheet.iter_rows(min_col=1, max_col=1, min_row=2, values_only=True):
            asin = row[0]  # Get the value from the first column
            if asin:  # Skip empty rows
                asins.append(str(asin))

        return asins
    except Exception as e:
        logging.error(f"Error reading ASINs from file {file_path}: {e}")
        return []

@app.route('/')
def home():
    return render_template('index.html')

@app.route('/process-asins', methods=['POST'])
def process_asins():
    if 'file' not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No selected file"}), 400

    if not file.filename.endswith('.xlsx'):
        return jsonify({"error": "Invalid file format. Only .xlsx allowed."}), 400

    file_path = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(file.filename))
    file.save(file_path)

    try:
        # Extract ASINs from the uploaded file
        global ASINS
        ASINS = extract_asins_from_file(file_path)

        # Scrape sellers and get the data
        seller_data = asyncio.run(scrape_sellers())
        
        # Create CSV string
        csv_string = create_csv_string(seller_data)
        
        # Create response with CSV file
        response = make_response(csv_string)
        response.headers['Content-Disposition'] = 'attachment; filename=amazon_sellers.csv'
        response.headers['Content-Type'] = 'text/csv'

        return response
    finally:
        # Ensure the uploaded file is deleted
        if os.path.exists(file_path):
            os.remove(file_path)

        # Reset global state
        ASINS.clear()
        cache.clear()

if __name__ == "__main__":
    port = int(os.getenv("PORT", 10000))  # Use the PORT environment variable or default to 10000
    app.run(host="0.0.0.0", port=port, debug=True)
